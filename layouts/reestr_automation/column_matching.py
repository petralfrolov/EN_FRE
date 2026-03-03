"""
Модуль сопоставления столбцов и листов для вкладки 'Автоматизация реестра'.
Нечёткий поиск (difflib), ручное подтверждение, сохранение выборов.
"""

import difflib
import logging

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

import database as db
from .constants import (
    DEFAULT_SHEET_MAPPING, FEED_TYPE_LABELS,
    DICT_RENAME, FUZZY_MATCH_CUTOFF,
)

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
#  Вспомогательные функции
# ═══════════════════════════════════════════════════════════════

def _get_user_settings() -> dict:
    username = st.session_state.get("username")
    if not username:
        return {}
    return db.get_user_settings(username)


def _save_user_settings(settings: dict):
    username = st.session_state.get("username")
    if username:
        db.update_user_settings(username, settings)


# ═══════════════════════════════════════════════════════════════
#  Сопоставление листов
# ═══════════════════════════════════════════════════════════════

def _get_actual_sheets(excel_path: str) -> list[str]:
    """Возвращает список листов из файла реестра."""
    try:
        wb = load_workbook(excel_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        st.error(f"Ошибка чтения реестра: {e}")
        return []


def get_sheet_mapping() -> dict | None:
    """
    Возвращает актуальный маппинг листов.
    Берёт из настроек пользователя, если сохранён.
    Иначе возвращает DEFAULT_SHEET_MAPPING.
    """
    settings = _get_user_settings()
    saved = settings.get("sheet_mapping")
    if saved and isinstance(saved, dict):
        return saved
    return dict(DEFAULT_SHEET_MAPPING)


def validate_sheet_mapping(excel_path: str) -> tuple[bool, list[str]]:
    """
    Проверяет, что все листы из текущего маппинга существуют в реестре.
    Возвращает (is_valid, actual_sheets).
    """
    actual_sheets = _get_actual_sheets(excel_path)
    if not actual_sheets:
        return False, []

    mapping = get_sheet_mapping()
    for feed_type, sheet_name in mapping.items():
        if sheet_name not in actual_sheets:
            return False, actual_sheets

    return True, actual_sheets


def render_sheet_mapping_ui(excel_path: str) -> bool:
    """
    Отображает UI для сопоставления типов кормов с листами реестра.
    Возвращает True если маппинг утверждён и сохранён.
    """
    actual_sheets = _get_actual_sheets(excel_path)
    if not actual_sheets:
        return False

    st.warning("⚠️ Листы в файле реестра не совпадают с ожидаемыми. Сопоставьте типы кормов с листами вручную.")

    current_mapping = get_sheet_mapping()
    new_mapping = {}

    for feed_type, label in FEED_TYPE_LABELS.items():
        current_sheet = current_mapping.get(feed_type, "")
        # Если текущее значение есть в списке — ставим его, иначе первый
        if current_sheet in actual_sheets:
            default_idx = actual_sheets.index(current_sheet)
        else:
            # Пытаемся найти похожий лист
            close = difflib.get_close_matches(
                DEFAULT_SHEET_MAPPING.get(feed_type, ""),
                actual_sheets, n=1, cutoff=0.4
            )
            default_idx = actual_sheets.index(close[0]) if close else 0

        selected = st.selectbox(
            f"**{label}** →",
            options=actual_sheets,
            index=default_idx,
            key=f"sheet_map_{feed_type}"
        )
        new_mapping[feed_type] = selected

    if st.button("✅ Сохранить маппинг листов", type="primary"):
        settings = _get_user_settings()
        settings["sheet_mapping"] = new_mapping
        _save_user_settings(settings)
        st.success("Маппинг листов сохранён!")
        st.rerun()
        return True

    return False


# ═══════════════════════════════════════════════════════════════
#  Сопоставление столбцов
# ═══════════════════════════════════════════════════════════════

def get_registry_headers(excel_path: str, sheet_name: str) -> list[str]:
    """Читает заголовки (строка 2) из листа реестра."""
    try:
        wb = load_workbook(excel_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        headers = [c.value for c in list(ws.iter_rows(min_row=2, max_row=2))[0] if c.value]
        wb.close()
        return headers
    except Exception as e:
        logger.error(f"Ошибка чтения заголовков: {e}")
        return []


def _get_saved_column_mappings(feed_type: str) -> dict:
    """Возвращает сохранённые маппинги столбцов для данного типа корма."""
    settings = _get_user_settings()
    all_mappings = settings.get("column_mappings", {})
    return all_mappings.get(feed_type, {})


def _save_column_mappings(feed_type: str, mappings: dict):
    """Сохраняет маппинги столбцов для данного типа корма."""
    settings = _get_user_settings()
    if "column_mappings" not in settings:
        settings["column_mappings"] = {}
    settings["column_mappings"][feed_type] = mappings
    _save_user_settings(settings)


def match_columns(
    source_cols: list[str],
    registry_cols: list[str],
    saved_mappings: dict,
) -> dict:
    """
    Сопоставляет столбцы файла со столбцами реестра.

    Возвращает dict:
      {source_col: {
          "match": str | None,     # найденное совпадение
          "source": str,           # откуда: "exact", "rename", "saved", "fuzzy", None
          "candidates": list[str], # кандидаты для ручного выбора
      }}
    """
    result = {}
    # Инвертируем DICT_RENAME: значение_реестра -> ключ_файла
    rename_forward = dict(DICT_RENAME)  # файл -> реестр
    registry_set = set(registry_cols)

    for col in source_cols:
        # 1. Точное совпадение
        if col in registry_set:
            result[col] = {"match": col, "source": "exact", "candidates": []}
            continue

        # 2. Совпадение через DICT_RENAME
        renamed = rename_forward.get(col)
        if renamed and renamed in registry_set:
            result[col] = {"match": renamed, "source": "rename", "candidates": []}
            continue

        # 3. Сохранённые маппинги пользователя
        saved_match = saved_mappings.get(col)
        if saved_match and saved_match in registry_set:
            result[col] = {"match": saved_match, "source": "saved", "candidates": []}
            continue

        # 4. Нечёткий поиск
        close = difflib.get_close_matches(
            col, registry_cols, n=5, cutoff=FUZZY_MATCH_CUTOFF
        )
        if close:
            result[col] = {"match": close[0], "source": "fuzzy", "candidates": close}
        else:
            result[col] = {"match": None, "source": None, "candidates": []}

    return result


def apply_column_mapping(df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """
    Переименовывает столбцы DataFrame по утверждённому маппингу.
    mapping: {source_col: target_col_or_None}
    """
    rename_dict = {}
    for src, tgt in mapping.items():
        if tgt and src != tgt and src in df.columns:
            rename_dict[src] = tgt
    return df.rename(columns=rename_dict)


def render_column_matching_ui(
    df: pd.DataFrame,
    feed_type: str,
    excel_path: str,
    sheet_name: str,
) -> tuple[pd.DataFrame | None, bool]:
    """
    Показывает UI сопоставления столбцов.
    Возвращает (df_с_переименованными_столбцами, all_resolved).

    all_resolved=True — можно записывать в реестр.
    """
    registry_cols = get_registry_headers(excel_path, sheet_name)
    if not registry_cols:
        st.error(f"Не удалось прочитать заголовки из листа '{sheet_name}'")
        return None, False

    saved_mappings = _get_saved_column_mappings(feed_type)
    source_cols = list(df.columns)
    match_result = match_columns(source_cols, registry_cols, saved_mappings)

    # Разделяем: автоматически сопоставленные и проблемные
    auto_matched = {}
    needs_review = {}
    for col, info in match_result.items():
        if info["source"] in ("exact", "rename", "saved"):
            auto_matched[col] = info["match"]
        else:
            needs_review[col] = info

    # Инициализируем session_state для маппинга
    state_key = f"col_mapping_{feed_type}"
    if state_key not in st.session_state:
        st.session_state[state_key] = {}

    # Объединяем auto + уже утверждённые
    approved = dict(auto_matched)
    approved.update(st.session_state[state_key])

    # Фильтруем: что ещё не утверждено
    pending = {
        col: info for col, info in needs_review.items()
        if col not in st.session_state[state_key]
    }

    if pending:
        st.warning(f"⚠️ {len(pending)} столбцов требуют ручного сопоставления:")

        h1, h2 = st.columns([2, 3])
        h1.caption("**Столбец из файла**")
        h2.caption("**Столбец реестра (выберите соответствие)**")

        taken_registry_cols = {v for v in approved.values() if v is not None}
        available_registry_cols = [c for c in registry_cols if c not in taken_registry_cols]

        items_to_approve = {}

        for i, (col, info) in enumerate(pending.items()):
            c1, c2 = st.columns([2, 3])
            with c1:
                st.text(col)
            with c2:
                valid_candidates = [c for c in info["candidates"] if c not in taken_registry_cols]
                options = ["— Пропустить —"] + valid_candidates + available_registry_cols
                # Убираем дубликаты, сохраняя порядок
                seen = set()
                unique_options = []
                for o in options:
                    if o not in seen:
                        seen.add(o)
                        unique_options.append(o)

                selected = st.selectbox(
                    "Варианты:",
                    options=unique_options,
                    index=0,
                    key=f"colmap_{feed_type}_{i}_{col}",
                    label_visibility="collapsed",
                )
                items_to_approve[col] = selected

        if st.button("✅ Принять выбранные", key=f"colmap_btn_accept_all_{feed_type}", type="primary"):
            new_saved = dict(saved_mappings)
            for col, selected in items_to_approve.items():
                if selected == "— Пропустить —":
                    st.session_state[state_key][col] = None
                else:
                    st.session_state[state_key][col] = selected
                    new_saved[col] = selected

            _save_column_mappings(feed_type, new_saved)
            st.rerun()

        return None, False

    # Все столбцы сопоставлены — применяем маппинг
    final_mapping = {}
    for col in source_cols:
        if col in approved and approved[col]:
            final_mapping[col] = approved[col]

    df_mapped = apply_column_mapping(df, final_mapping)

    # Предупреждение о незаполненных столбцах реестра
    mapped_targets = set(final_mapping.values()) | set(
        col for col in source_cols if col in set(registry_cols)
    )
    unmapped_registry = [c for c in registry_cols if c not in mapped_targets]
    if unmapped_registry:
        with st.expander(f"ℹ️ {len(unmapped_registry)} столбцов реестра останутся пустыми"):
            st.write(", ".join(unmapped_registry[:20]))
            if len(unmapped_registry) > 20:
                st.write(f"...и ещё {len(unmapped_registry) - 20}")

    return df_mapped, True
