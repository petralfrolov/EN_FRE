"""
UI-рендер вкладки 'Автоматизация реестра'.
"""

import streamlit as st
import pandas as pd
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment

import database as db
from config import EXCEL_FILE_PATH, REESTR_NORM_FILE_PATH
from modules.codebook import load_codebook

from .constants import DICT_RENAME, THIN_BORDER, SHEETS_TO_COLOR
from .processing import process_lab_file, process_ro_tap_pair, read_lab_columns
from .norms import load_all_norm_rules, colorize_new_rows
from .column_matching import (
    get_sheet_mapping, validate_sheet_mapping,
    render_sheet_mapping_ui, render_column_matching_ui,
)

logger = logging.getLogger(__name__)


def log_action(action: str, details: str = "", level: str = "INFO"):
    """Логирует действие пользователя с контекстом для process mining."""
    extra = {
        'session_id': st.session_state.get('session_id', '-'),
        'user': st.session_state.get('username', '-'),
        'action': action,
        'details': details
    }
    if level == "WARNING":
        logger.warning("", extra=extra)
    elif level == "ERROR":
        logger.error("", extra=extra)
    else:
        logger.info("", extra=extra)


def get_user_reestr_norm_path() -> str:
    """Возвращает путь к файлу границ питательности (пользовательский или дефолтный).
    Если пользовательский путь недоступен, возвращает дефолтный.
    """
    username = st.session_state.get("username")
    if username:
        settings = db.get_user_settings(username)
        user_path = settings.get("reestr_norm_file_path")
        if user_path:
            if os.path.exists(user_path):
                return user_path
            else:
                st.warning(f"⚠️ Путь к границам питательности недоступен: '{user_path}'. Используется стандартный.")
    return REESTR_NORM_FILE_PATH


def get_user_excel_path() -> str:
    """Возвращает путь к файлу реестра (пользовательский или дефолтный).
    Если пользовательский путь недоступен, возвращает дефолтный.
    """
    username = st.session_state.get("username")
    if username:
        settings = db.get_user_settings(username)
        user_path = settings.get("excel_file_path")
        if user_path:
            if os.path.exists(user_path):
                return user_path
            else:
                st.warning(f"⚠️ Путь к реестру недоступен: '{user_path}'. Используется стандартный.")
    return EXCEL_FILE_PATH


def append_to_registry(df_new: pd.DataFrame, sheet_name: str, excel_path: str = None) -> tuple[bool, str]:
    """
    Дописать данные в указанный лист реестра.
    Возвращает (success, message).
    """
    if df_new.empty:
        return False, "Нет данных для записи"

    # Получаем путь к реестру (если не передан явно временный)
    if excel_path is None:
        excel_path = get_user_excel_path()

    # Применяем переименование
    df_new = df_new.rename(columns=DICT_RENAME)

    try:
        wb = load_workbook(excel_path)
    except FileNotFoundError:
        return False, f"Файл реестра не найден: {excel_path}"
    except Exception as e:
        return False, f"Ошибка открытия реестра: {e}"

    if sheet_name not in wb.sheetnames:
        return False, f"Лист '{sheet_name}' отсутствует в реестре"

    ws = wb[sheet_name]

    header_row_idx = 2
    last_old_row = ws.max_row

    headers = [c.value for c in ws[header_row_idx]]
    col_name_to_idx = {name: i + 1 for i, name in enumerate(headers) if name}

    template_row = last_old_row if last_old_row >= header_row_idx + 1 else None

    rows_written = 0

    for i, (_, row) in enumerate(df_new.iterrows(), start=1):
        excel_row = last_old_row + i

        for col_name, value in row.items():
            if col_name not in col_name_to_idx:
                continue

            col_idx = col_name_to_idx[col_name]
            cell = ws.cell(row=excel_row, column=col_idx, value=value)

            if template_row is not None:
                tpl = ws.cell(row=template_row, column=col_idx)
                cell.number_format = tpl.number_format

                if tpl.alignment and tpl.alignment.wrap_text:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        horizontal=tpl.alignment.horizontal,
                        vertical=tpl.alignment.vertical
                    )

            cell.border = THIN_BORDER
            cell.fill = PatternFill(fill_type=None)

        rows_written += 1

    # --- Покраска ячеек для листов из SHEETS_TO_COLOR ---
    first_new_excel_row = last_old_row + 1
    if sheet_name in SHEETS_TO_COLOR:
        try:
            norm_rules_by_sheet = load_all_norm_rules()
            colorize_new_rows(ws, df_new.reset_index(drop=True), first_new_excel_row, norm_rules_by_sheet)
        except Exception as e:
            # Ошибка покраски не критична, продолжаем сохранение
            pass

    try:
        wb.save(excel_path)
        return True, f"Добавлено {rows_written} строк в лист '{sheet_name}'"
    except PermissionError:
        return False, f"Ошибка доступа! Закройте файл реестра и попробуйте снова."
    except Exception as e:
        return False, f"Не удалось сохранить: {e}"


# ═══════════════════════════════════════════════════════════════
#  Общая функция рендера вкладки типа корма
# ═══════════════════════════════════════════════════════════════

def _get_sample_col_setting() -> str | None:
    """Получить сохранённый столбец 'номер образца' из настроек пользователя."""
    username = st.session_state.get("username")
    if not username:
        return None
    settings = db.get_user_settings(username)
    return settings.get("sample_col_name")


def _save_sample_col_setting(col_name: str):
    """Сохранить выбранный столбец 'номер образца' в настройки пользователя."""
    username = st.session_state.get("username")
    if not username:
        return
    settings = db.get_user_settings(username)
    settings["sample_col_name"] = col_name
    db.update_user_settings(username, settings)


def _render_sample_col_selector(
    columns: list[str],
    feed_type: str,
    uploaded_file_name: str,
) -> str | None:
    """
    Показывает UI выбора столбца 'номер образца'.
    Возвращает выбранный столбец или None, если ещё не выбран.
    """
    state_key = f"sample_col_confirmed_{feed_type}"
    state_file_key = f"sample_col_file_{feed_type}"

    # Сброс при смене файла
    if st.session_state.get(state_file_key) != uploaded_file_name:
        st.session_state.pop(state_key, None)
        st.session_state[state_file_key] = uploaded_file_name

    # Уже подтверждён для этого файла
    if state_key in st.session_state:
        return st.session_state[state_key]

    # Попробовать использовать сохранённый из настроек
    saved = _get_sample_col_setting()
    if saved and saved in columns:
        st.session_state[state_key] = saved
        return saved

    # Нужно выбрать — показываем UI
    if saved and saved not in columns:
        st.warning(
            f"⚠️ Ранее выбранный столбец **'{saved}'** не найден в этом файле. "
            f"Выберите столбец заново."
        )

    # Подсказка: попробовать угадать
    default_idx = 0
    for i, c in enumerate(columns):
        low = str(c).lower()
        if "образ" in low or "sample" in low:
            default_idx = i
            break

    chosen = st.selectbox(
        "Выберите столбец с номером образца",
        options=columns,
        index=default_idx,
        key=f"sample_col_select_{feed_type}",
    )

    if st.button("✅ Подтвердить столбец", key=f"confirm_sample_col_{feed_type}"):
        _save_sample_col_setting(chosen)
        st.session_state[state_key] = chosen
        st.rerun()

    return None  # Ещё не подтверждено — блокируем дальнейшую обработку


def _render_feed_tab(
    feed_type: str,
    label: str,
    file_key: str,
    df_key: str,
    fname_key: str,
    geo_map, culture_map, feed_map,
    excel_path: str,
    sheet_mapping: dict,
    temp_registry_name: str = None,
):
    """
    Общий рендер для вкладок основных кормов, сена, зелёной массы, HMC.
    """
    st.subheader(label)
    uploaded = st.file_uploader(
        "Загрузите файл анализа",
        type=["xlsx", "xls"],
        key=file_key,
    )

    if not uploaded:
        return

    # --- Выбор столбца «номер образца» ---
    cols_key = f"raw_columns_{feed_type}"
    if cols_key not in st.session_state or st.session_state.get(fname_key) != uploaded.name:
        raw_columns = read_lab_columns(uploaded)
        uploaded.seek(0)  # сбрасываем позицию после чтения заголовков
        st.session_state[cols_key] = raw_columns

    raw_columns = st.session_state[cols_key]
    if not raw_columns:
        st.error("Не удалось прочитать заголовки файла.")
        return

    sample_col = _render_sample_col_selector(raw_columns, feed_type, uploaded.name)
    if sample_col is None:
        st.info("Выберите и подтвердите столбец с номером образца для продолжения.")
        return

    # Обработка файла
    if df_key not in st.session_state or st.session_state.get(fname_key) != uploaded.name:
        with st.spinner("Обработка файла..."):
            df = process_lab_file(uploaded, feed_type, geo_map, culture_map, feed_map, sample_col=sample_col)
        st.session_state[df_key] = df
        st.session_state[fname_key] = uploaded.name
        # Сбросить предыдущий маппинг столбцов
        col_map_key = f"col_mapping_{feed_type}"
        if col_map_key in st.session_state:
            del st.session_state[col_map_key]

    df = st.session_state[df_key]
    if df.empty:
        return

    st.success(f"Обработано строк: {len(df)}")

    # Редактируемая таблица
    edited_df = st.data_editor(
        df,
        width='stretch',
        height=400,
        num_rows="dynamic",
        key=f"editor_{feed_type}",
    )
    st.session_state[df_key] = edited_df

    # --- Сопоставление столбцов ---
    sheet_name = sheet_mapping.get(feed_type, "")
    df_mapped, all_resolved = render_column_matching_ui(
        edited_df, feed_type, excel_path, sheet_name,
    )

    # Кнопка записи
    if st.button(
        "Добавить в реестр",
        key=f"btn_{feed_type}",
        type="primary",
        disabled=not all_resolved,
    ):
        df_to_write = df_mapped if df_mapped is not None else edited_df
        with st.spinner("Добавление в реестр..."):
            success, msg = append_to_registry(df_to_write, sheet_name, excel_path)
        if success:
            log_action("REESTR_APPEND", f"type={feed_type}, rows={len(df_to_write)}")
            st.success(msg)
            # Для кастомного реестра — предложить скачать обновлённый файл
            if temp_registry_name:
                try:
                    with open(excel_path, "rb") as f:
                        st.download_button(
                            "📥 Скачать обновлённый реестр",
                            data=f,
                            file_name=temp_registry_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_after_{feed_type}",
                        )
                except Exception as e:
                    st.warning(f"Не удалось подготовить файл для скачивания: {e}")
        else:
            log_action("REESTR_APPEND_ERROR", f"type={feed_type}, error={msg}", level="ERROR")
            st.error(msg)


# ═══════════════════════════════════════════════════════════════
#  Основная функция рендера
# ═══════════════════════════════════════════════════════════════

def render_reestr_automation():
    """
    Отрисовывает вкладку 'Автоматизация реестра'.
    """
    st.subheader("Автоматизация реестра кормов")

    # Загружаем справочники
    geo_map, culture_map, feed_map = load_codebook()

    if geo_map.empty:
        st.error("Справочники не загружены. Проверьте путь к файлу Кодировка.xlsx")
        return

    # --- Загрузка стороннего реестра ---
    with st.expander("Загрузить сторонний реестр (опционально)"):
        temp_registry = st.file_uploader(
            "Загрузите Excel-файл реестра, если хотите работать с ним вместо базы",
            type=["xlsx", "xls"],
            key="temp_registry_upload"
        )

    excel_path = get_user_excel_path()
    using_temp = False
    temp_registry_name = None

    if temp_registry:
        temp_path = "temp_reestr_upload.xlsx"
        try:
            with open(temp_path, "wb") as f:
                f.write(temp_registry.getvalue())
            excel_path = temp_path
            using_temp = True
            temp_registry_name = temp_registry.name
            st.info("Используется загруженный сторонний реестр. После добавления данных появится кнопка скачивания.")
        except Exception as e:
            st.error(f"Ошибка сохранения временного реестра: {e}")

    if not using_temp:
        st.info(f"Реестр (база): `{excel_path}`")

    # ── Проверка маппинга листов ──
    is_valid, actual_sheets = validate_sheet_mapping(excel_path)
    if not is_valid:
        render_sheet_mapping_ui(excel_path)
        return  # Блокируем дальнейшую работу

    sheet_mapping = get_sheet_mapping()

    # --- Вкладки для разных типов файлов ---
    tabs = st.tabs([
        "Основные корма",
        "Сено, солома",
        "Зелёная масса",
        "Карнаж/плющ.зерно",
        "RoTap"
    ])

    # --- Основные корма ---
    with tabs[0]:
        _render_feed_tab(
            feed_type="main_feeds",
            label="Основные корма (силос)",
            file_key="file_main_feeds",
            df_key="df_main_feeds",
            fname_key="main_file_name",
            geo_map=geo_map, culture_map=culture_map, feed_map=feed_map,
            excel_path=excel_path,
            sheet_mapping=sheet_mapping,
            temp_registry_name=temp_registry_name,
        )

    # --- Сено, солома ---
    with tabs[1]:
        _render_feed_tab(
            feed_type="hay_straw",
            label="Сено, солома",
            file_key="file_hay_straw",
            df_key="df_hay_straw",
            fname_key="hay_file_name",
            geo_map=geo_map, culture_map=culture_map, feed_map=feed_map,
            excel_path=excel_path,
            sheet_mapping=sheet_mapping,
            temp_registry_name=temp_registry_name,
        )

    # --- Зелёная масса ---
    with tabs[2]:
        _render_feed_tab(
            feed_type="green",
            label="Зелёная масса",
            file_key="file_green",
            df_key="df_green",
            fname_key="green_file_name",
            geo_map=geo_map, culture_map=culture_map, feed_map=feed_map,
            excel_path=excel_path,
            sheet_mapping=sheet_mapping,
            temp_registry_name=temp_registry_name,
        )

    # --- Карнаж/плющ.зерно ---
    with tabs[3]:
        _render_feed_tab(
            feed_type="hmc",
            label="Карнаж и плющеное зерно (HMC)",
            file_key="file_hmc",
            df_key="df_hmc",
            fname_key="hmc_file_name",
            geo_map=geo_map, culture_map=culture_map, feed_map=feed_map,
            excel_path=excel_path,
            sheet_mapping=sheet_mapping,
            temp_registry_name=temp_registry_name,
        )

    # --- RoTap ---
    with tabs[4]:
        st.subheader("RoTap анализ")
        st.markdown("Загрузите **два файла**: лабораторный анализ и результаты Ro-Tap.")

        col1, col2 = st.columns(2)

        with col1:
            file_lab = st.file_uploader(
                "Лабораторный файл (corn silage ena...)",
                type=["xlsx", "xls"],
                key="file_rotap_lab"
            )

        with col2:
            file_rt = st.file_uploader(
                "Rezultatyi_Ro_Tap.xlsx",
                type=["xlsx", "xls"],
                key="file_rotap_results"
            )

        if file_lab and file_rt:
            file_key = f"{file_lab.name}_{file_rt.name}"

            # --- Выбор столбца «номер образца» для RoTap ---
            rt_cols_key = "raw_columns_rotap"
            if rt_cols_key not in st.session_state or st.session_state.get("rotap_file_key") != file_key:
                raw_cols = read_lab_columns(file_lab, usecols="B:AW")
                file_lab.seek(0)
                st.session_state[rt_cols_key] = raw_cols

            raw_cols = st.session_state[rt_cols_key]
            if not raw_cols:
                st.error("Не удалось прочитать заголовки лабораторного файла.")
                return

            sample_col = _render_sample_col_selector(raw_cols, "ro_tap", file_lab.name)
            if sample_col is None:
                st.info("Выберите и подтвердите столбец с номером образца для продолжения.")
                return

            if "df_rotap" not in st.session_state or st.session_state.get("rotap_file_key") != file_key:
                with st.spinner("Обработка файлов..."):
                    df = process_ro_tap_pair(file_lab, file_rt, geo_map, culture_map, feed_map, sample_col=sample_col)
                st.session_state.df_rotap = df
                st.session_state.rotap_file_key = file_key
                # Сбросить предыдущий маппинг столбцов
                col_map_key = "col_mapping_ro_tap"
                if col_map_key in st.session_state:
                    del st.session_state[col_map_key]

            df = st.session_state.df_rotap

            if not df.empty:
                st.success(f"Обработано строк: {len(df)}")

                edited_df = st.data_editor(
                    df,
                    width='stretch',
                    height=400,
                    num_rows="dynamic",
                    key="editor_rotap"
                )
                st.session_state.df_rotap = edited_df

                # --- Сопоставление столбцов ---
                rt_sheet = sheet_mapping.get("ro_tap", "")
                df_mapped, all_resolved = render_column_matching_ui(
                    edited_df, "ro_tap", excel_path, rt_sheet,
                )

                if st.button(
                    "Добавить в реестр",
                    key="btn_rotap",
                    type="primary",
                    disabled=not all_resolved,
                ):
                    df_to_write = df_mapped if df_mapped is not None else edited_df
                    with st.spinner("Добавление в реестр..."):
                        success, msg = append_to_registry(df_to_write, rt_sheet, excel_path)
                    if success:
                        log_action("REESTR_APPEND", f"type=ro_tap, rows={len(df_to_write)}")
                        st.success(msg)
                        # Для кастомного реестра — предложить скачать обновлённый файл
                        if temp_registry_name:
                            try:
                                with open(excel_path, "rb") as f:
                                    st.download_button(
                                        "📥 Скачать обновлённый реестр",
                                        data=f,
                                        file_name=temp_registry_name,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key="dl_after_rotap",
                                    )
                            except Exception as e:
                                st.warning(f"Не удалось подготовить файл для скачивания: {e}")
                    else:
                        log_action("REESTR_APPEND_ERROR", f"type=ro_tap, error={msg}", level="ERROR")
                        st.error(msg)
        elif file_lab or file_rt:
            st.warning("Загрузите оба файла для обработки RoTap")
